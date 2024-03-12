import os
import openpyxl
import module

DATA_DIR = './data/'
RESULTS_DIR = './results/'

filename = module.get_file()
data_path = os.path.join(DATA_DIR, filename)
wb = openpyxl.load_workbook(data_path)
ws = wb.active

# extract all data first (row 6 onwards)
data = [value for value in ws.iter_rows(
            min_row=6, max_row=ws.max_row, min_col=4, max_col=ws.max_column, 
            values_only=True
        )
    ]

# extract headers and answer key out
headers = data.pop(0) + ('Score',)
answers = data.pop(0)[2:]

# loop through each row to collect responses of respondents
respondents = []

for i in range(len(data)):

    respondent = module.Respondent(data[i][0], data[i][1], data[i][2:])
    respondent.check_response(answers)
    respondents.append(respondent)

# write results into new sheet
wb.create_sheet('Results')
ws = wb['Results']

# write the headers in first
for i in range(len(headers)):

    ws.cell(row=1, column=i + 1, value=headers[i])

# write the results for each person
curr_row = 2

for respondent in respondents:

    module.write_respondent_data(ws, respondent, curr_row)
    curr_row += 1

result_path = os.path.join(RESULTS_DIR, filename)
wb.save(result_path)
os.remove(data_path)
