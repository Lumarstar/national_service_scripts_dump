import os
import csv
import openpyxl
import openpyxl.styles


class Respondent():

    def __init__(self, company, name, responses):

        self._company = company
        self._name = name
        self._responses = responses
        self._score = 0
        self._wrong_responses = []
    
    @property
    def company(self):

        return self._company

    @company.setter
    def company(self, company):

        self._company = company

    @property
    def name(self):

        return self._name

    @name.setter
    def name(self, name):

        self._name = name
        
    @property
    def responses(self):

        return self._responses

    @responses.setter
    def responses(self, responses):

        self._responses = responses

    @property
    def score(self):

        return self._score

    @score.setter
    def score(self, score):

        self._score = score

    @property
    def wrong_responses(self):

        return self._wrong_responses

    @wrong_responses.setter
    def wrong_responses(self, wrong_responses):

        self._wrong_responses = wrong_responses

    def check_response(self, answers):

        """
        method to check responses given a list `answers`, and identify questions that were answered incorrectly,
        as well as calculate the score of the respondent 
        """

        for i in range(len(answers)):

            if answers[i] == self.responses[i]:

                self.score += 1

            else:

                self.wrong_responses.append(i)


def get_file():

    """A function that formats a file into `.xlsx` format and returns the filename"""

    dir_path = os.curdir + '/data/'
    filename = list(os.listdir(dir_path))[0]
    filepath = dir_path + filename
    
    if filepath.endswith('.csv'):

        wb = openpyxl.Workbook()
        ws = wb.active

        with open(filepath) as f:

            reader = csv.reader(f, delimiter=',')
            
            for row in reader:

                ws.append(row)

        os.remove(filepath)

        filename = filename[:-3] + 'xlsx'
        filepath = dir_path + filename
        wb.save(filepath)

    return filename


def fill_cell(ws, row, col):

    """Fills a cell specified by its row and column with a solid colour."""

    highlight_fill = openpyxl.styles.PatternFill(
        fill_type = 'solid',
        start_color = 'FFCCCB',
        end_color = 'FFCCCB'
    )

    ws.cell(row=row, column=col).fill = highlight_fill


def write_respondent_data(ws, respondent, row):

    """Writes data from `respondent` in `ws` at row `row`."""

    ws.cell(row=row, column=1, value=respondent.company)
    ws.cell(row=row, column=2, value=respondent.name)

    for i, response in enumerate(respondent.responses):
        ws.cell(row=row, column=i + 3, value=response)

        if i in respondent.wrong_responses:
            fill_cell(ws, row, i + 3)

    ws.cell(row=row, column=ws.max_column, value=respondent.score)